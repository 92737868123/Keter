import discord
import time
import psutil
import os
import asyncio
import openpyxl
import random
import math
import numpy as np
import datetime
import pandas_datareader.data as web
import matplotlib.pyplot as plt

from datetime import datetime
from discord.ext import commands
from evs import default, permissions

userlib = "./lib/economy/users/"
stocklib = "./lib/economy/stocks/"


class economy_ko(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.config = default.get("config.json")
        self.process = psutil.Process(os.getpid())
        # 폴더생성
        if os.path.isdir("./lib/economy/users"):
            print("user folder exist")
        else:
            os.makedirs("./lib/economy/users")

        if os.path.isdir("./lib/economy/stocks"):
            print("stocks folder exist")
        else:
            os.makedirs("./lib/economy/stocks")

    # 메시지당 돈
    @commands.Cog.listener()
    async def on_message(self, ctx):
        if ctx.guild.id == 749595288280498188:
            if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
                randomnum = random.randrange(1, 3)
                wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
                ws = wb.active
                suvmoney = int(ws.cell(row=1, column=2).value)
                suvmoney = suvmoney + randomnum
                ws.cell(row=1, column=2).value = str(suvmoney)
                wb.save(userlib + str(ctx.author.id) + ".xlsx")
                wb.close()

    # 참여
    @commands.command()
    async def 참여(self, ctx):

        embed = discord.Embed(title="케테르 경제", description="케테르 경제에 참여하시겠습니까?", color=0xeff0f1)
        embed.set_thumbnail(
            url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
        msg = await ctx.send(embed=embed)

        def reaction_check_(m):
            if m.message_id == msg.id and m.user_id == ctx.author.id and str(m.emoji) == "✅":
                return True
            return False

        try:
            await msg.add_reaction("✅")
            await self.bot.wait_for('raw_reaction_add', timeout=10.0, check=reaction_check_)
            if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
                embed = discord.Embed(title="케테르 경제", description="이미 참여하셨습니다.", color=0xeff0f1)
                await ctx.send(embed=embed)
            else:
                embed = discord.Embed(title="케테르 경제",
                                      description="새로 오셨군요? " + str(ctx.author.name) + "님을 위한 파일들을 생성중이에요!",
                                      color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
                await ctx.send(embed=embed)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.cell(row=1, column=1).value = "Hello World"  #:)
                ws.cell(row=1, column=2).value = "8600000"  # money
                ws.cell(row=1, column=3).value = "0"  # pres
                ws.cell(row=1, column=4).value = "-"  # rank
                ws.cell(row=2, column=1).value = "None"  # status
                ws.cell(row=2, column=2).value = "0"  # perfect
                ws.cell(row=2, column=3).value = "0"  # great
                ws.cell(row=2, column=4).value = "0"  # good
                ws.cell(row=2, column=5).value = "0"  # bad
                ws.cell(row=3, column=1).value = "0"  # tsucc
                ws.cell(row=3, column=2).value = "0"  # tfail
                ws.cell(row=3, column=3).value = "0"  # fails
                ws.cell(row=4, column=1).value = "0"  # home count
                ws.cell(row=4, column=2).value = "[1]"  # title
                ws.cell(row=4, column=3).value = "1"  # header
                ws.cell(row=4, column=4).value = "1"  # tail
                ws.cell(row=5, column=1).value = "100"  # HP
                ws.cell(row=5, column=2).value = "100"  # STR
                ws.cell(row=5, column=3).value = "100"  # DEF
                ws.cell(row=5, column=4).value = "100"  # INT
                wb.save(userlib + str(ctx.author.id) + ".xlsx")
                wb.close()
                time.sleep(1)
                embed = discord.Embed(title="케테르 경제",
                                      description=str(ctx.author.name) + " 생성 완료!",
                                      color=0xeff0f1)
                await ctx.send(embed=embed)

        except asyncio.TimeoutError:
            await msg.delete()
            embed = discord.Embed(title="케테르 경제", description="서명하지 않으셨습니다. 다음 기회에..", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752690012369190942/DARK_KETER_1.png")
            await ctx.send(embed=embed)

        except discord.Forbidden:
            embed = discord.Embed(title="케테르 경제", description="케테르 경제에 참여하시겠습니까?", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await msg.edit(content=embed)

    @commands.command(aliases=['돈내놔', '돈줘'])
    async def 돈받기(self, ctx):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            num = random.randrange(1, 12)
            jackpot = random.random()
            if jackpot < 0.001:
                num = num * 10000
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            getmoney = ws.cell(row=1, column=2).value
            getmoney = int(getmoney) + int(num)
            ws.cell(row=1, column=2).value = str(getmoney)
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KET", description="<@" + str(ctx.author.id) + "> " + str(
                num) + "<:ket:753449741186105375>을 받았어요!", color=0xeff0f1)
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command()
    async def 돈(self, ctx):
        def keundon(value: int):
            value = int(value)
            if value < 0:
                return "변수는 음수값을 가질 수 없습니다."
            elif 0 <= value < 10000:
                return str(value)
            elif 10000 <= value < 100000000:
                return str(math.floor(value / 10000)) + "만 " + str(value - math.floor(value / 10000) * 10000)
            elif 100000000 <= value < 1000000000000:
                return str(math.floor(value / 100000000)) + "억 " + str(
                    math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(
                    value - math.floor(value / 10000) * 10000)
            elif 1000000000000 <= value < 10000000000000000:
                return str(math.floor(value / 1000000000000)) + "조 " + str(
                    math.floor(value / 100000000) - math.floor(value / 1000000000000) * 10000) + "억 " + str(
                    math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(
                    value - math.floor(value / 10000) * 10000)
            elif 10000000000000000 <= value < 100000000000000000000:
                return str(math.floor(value / 10000000000000000)) + "경 " + str(
                    math.floor(value / 1000000000000) - math.floor(value / 10000000000000000) * 10000) + "조 " + str(
                    math.floor(value / 100000000) - math.floor(value / 1000000000000) * 10000) + "억 " + str(
                    math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(
                    value - math.floor(value / 10000) * 10000)
            else:
                return "변수의 크기가 너무 큽니다."

        if (ctx.message.mentions.__len__() > 0):
            for user in ctx.message.mentions:
                if os.path.isfile(userlib + str(user.id) + ".xlsx"):
                    wb = openpyxl.load_workbook(userlib + str(user.id) + ".xlsx")
                    ws = wb.active
                    money = ws.cell(row=1, column=2).value
                    wb.close()
                    kundon = keundon(money)
                    embed = discord.Embed(title="KET", description="<@" + str(
                        user.id) + ">님은 " + kundon + "<:ket:753449741186105375>을 가지고 계십니다!", color=0xeff0f1)
                    await ctx.send(embed=embed)
                else:
                    embed = discord.Embed(title="NO", description="유저가 ``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
                    embed.set_thumbnail(
                        url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                    await ctx.send(embed=embed)
        else:
            if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
                wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
                ws = wb.active
                money = ws.cell(row=1, column=2).value
                wb.close()
                kundon = keundon(money)
                embed = discord.Embed(title="KET", description="<@" + str(
                    ctx.author.id) + "> " + kundon + "<:ket:753449741186105375>을 가지고 계십니다!", color=0xeff0f1)
                await ctx.send(embed=embed)
            else:
                embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)

    @commands.command(aliases=['프리스티지', '프레스티지', 'ㅎㅍ'])
    async def 호프(self, ctx):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            prestige = ws.cell(row=1, column=3).value
            wb.close()
            embed = discord.Embed(title="PRESTIGE", description="<@" + str(ctx.author.id) + "> " + str(
                prestige) + "<:pre:753458787465297993>을 가지고 계십니다!", color=0xeff0f1)
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['ㄷㅂ'])
    async def 도박(self, ctx, val: int):
        if val <= 0:
            embed = discord.Embed(title="NO", description="0 이하로는 베팅할 수 없어요.", color=0xeff0f1)
            await ctx.send(embed=embed)
            return None
        if val > 80000000000:
            embed = discord.Embed(title="NO", description="베팅금은 800억 을 초과할 수 없어요.", color=0xeff0f1)
            await ctx.send(embed=embed)
            return None
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            money = ws.cell(row=1, column=2).value
            if int(money) > val:
                discrim = random.random()
                if discrim < 0.02:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + 11 * val)
                    ws.cell(row=3, column=3).value = "0"
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "축하합니다! 대박이 나서 12배를 획득 하셨어요! 🎉\n획득량:" + str(
                        12 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
                elif 0.02 < discrim < 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + 2 * val)
                    ws.cell(row=3, column=3).value = "0"
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "축하합니다! 도박에 성공하셔서 3배를 획득 하셨어요! 🎉\n획득량:" + str(
                        3 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
                elif 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100 < discrim < 0.1 + math.sqrt(
                        int(ws.cell(row=3, column=3).value) * 100) / 50:
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) + val)
                    ws.cell(row=3, column=3).value = "0"
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "축하합니다! 도박에 성공하셔서 2배를 획득 하셨어요! 🎉\n획득량:" + str(
                        2 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
                else:
                    emj = "<:dar:754345236574109716>"
                    ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) - val)
                    ws.cell(row=3, column=3).value = str(int(ws.cell(row=3, column=3).value) + 1)
                    embed = discord.Embed(title="도박", description="<@" + str(
                        ctx.author.id) + "> " + "도박에 실패하여 돈을 잃으셨습니다. " + emj, color=0xeff0f1)
                wb.save(userlib + str(ctx.author.id) + ".xlsx")
                wb.close()
                await ctx.send(embed=embed)
            else:
                embed = discord.Embed(title="NO", description="보유하신 잔액보다 큰 금액을 베팅할 수는 없어요.", color=0xeff0f1)
                await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['ㅇㅇ'])
    async def 올인(self, ctx):
        if os.path.isfile(userlib + str(ctx.author.id) + ".xlsx"):
            wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
            ws = wb.active
            val = int(ws.cell(row=1, column=2).value)
            if val > 80000000000:
                embed = discord.Embed(title="NO", description="전재산이 800억을 초과하여 올인을 사용하실 수 없습니다.", color=0xeff0f1)
                await ctx.send(embed=embed)
                return None
            discrim = random.random()
            if discrim < 0.02:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) * 12)
                ws.cell(row=3, column=3).value = "0"
                embed = discord.Embed(title="올인", description="<@" + str(
                    ctx.author.id) + "> " + "축하합니다! 대박이 나서 12배를 획득 하셨어요! 🎉\n획득량:" + str(
                    12 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
            elif 0.02 < discrim < 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) * 3)
                ws.cell(row=3, column=3).value = "0"
                embed = discord.Embed(title="올인", description="<@" + str(
                    ctx.author.id) + "> " + "축하합니다! 올인에 성공하셔서 3배를 획득 하셨어요! 🎉\n획득량:" + str(
                    3 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
            elif 0.05 + math.sqrt(int(ws.cell(row=3, column=3).value) * 100) / 100 < discrim < 0.1 + math.sqrt(
                    int(ws.cell(row=3, column=3).value) * 100) / 50:
                ws.cell(row=1, column=2).value = str(int(ws.cell(row=1, column=2).value) * 2)
                ws.cell(row=3, column=3).value = "0"
                embed = discord.Embed(title="올인", description="<@" + str(
                    ctx.author.id) + "> " + "축하합니다! 올인에 성공하셔서 2배를 획득 하셨어요! 🎉\n획득량:" + str(
                    2 * val) + " <:ket:753449741186105375>", color=0xeff0f1)
            else:
                emj = "<:dar:754345236574109716>"
                ws.cell(row=1, column=2).value = "0"
                ws.cell(row=3, column=3).value = str(int(ws.cell(row=3, column=3).value) + 1)
                embed = discord.Embed(title="도박", description="올인에 실패하여 전재산을 잃으셨습니다. " + emj, color=0xeff0f1)
            wb.save(userlib + str(ctx.author.id) + ".xlsx")
            wb.close()
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="먼저 ``.참여``를 입력해서 케테르 경제에 참여해주세요!", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command()
    async def 송금(self, ctx, mention: str, valu: int):
        if (ctx.message.mentions.__len__() > 0):
            for user in ctx.message.mentions:
                if os.path.isfile(userlib + str(user.id) + ".xlsx"):
                    wb = openpyxl.load_workbook(userlib + str(ctx.author.id) + ".xlsx")
                    ws = wb.active
                    money = int(ws.cell(row=1, column=2).value)
                    if int(money) > valu:
                        wb2 = openpyxl.load_workbook(userlib + str(user.id) + ".xlsx")
                        ws2 = wb2.active
                        money2 = int(ws2.cell(row=1, column=2).value)
                        money2 = money2 + round(valu * 92 / 100)
                        ws2.cell(row=1, column=2).value = str(money2)
                        wb2.save(userlib + str(user.id) + ".xlsx")
                        wb2.close()
                        money = money - valu
                        ws.cell(row=1, column=2).value = str(money)
                        wb.save(userlib + str(ctx.author.id) + ".xlsx")
                        wb.close()
                        embed = discord.Embed(title="송금", description="<@" + str(ctx.author.id) + "> " + str(
                            round(valu * 92 / 100)) + " <:ket:753449741186105375>" + "송금 완료(세율 8%)", color=0xeff0f1)
                        await ctx.send(embed=embed)
                    else:
                        embed = discord.Embed(title="NO", description="보유하신 잔액보다 큰 금액을 송금할 수는 없어요.", color=0xeff0f1)
                        await ctx.send(embed=embed)
                else:
                    embed = discord.Embed(title="NO", description="유저가 ``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
                    embed.set_thumbnail(
                        url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                    await ctx.send(embed=embed)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 초기화(self, ctx):
        file_list = os.listdir(userlib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        for i in range(len(file_list)):
            wb = openpyxl.load_workbook(userlib + file_list[i])
            ws = wb.active
            if int(ws.cell(row=1, column=3).value) <= 1000:
                ws.cell(row=1, column=3).value = str(
                    int(ws.cell(row=1, column=3).value) + math.ceil(int(ws.cell(row=1, column=2).value) / 1000000000))
            else:
                ws.cell(row=1, column=3).value = str(round(int(ws.cell(row=1, column=3).value) / 2) + math.ceil(
                    int(ws.cell(row=1, column=2).value) / 1000000000))
            ws.cell(row=1, column=2).value = "8600000"
            wb.save(userlib + file_list[i])
            wb.close()
        embed = discord.Embed(title="Admin", description="초기화 완료", color=0xeff0f1)
        await ctx.send(embed=embed)

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 돈추가(self, ctx, mention: str, value: int):
        if (ctx.message.mentions.__len__() > 0):
            for user in ctx.message.mentions:
                if os.path.isfile(userlib + str(user.id) + ".xlsx"):
                    wb = openpyxl.load_workbook(userlib + str(user.id) + ".xlsx")
                    ws = wb.active
                    money = ws.cell(row=1, column=2).value
                    money = int(money) + value
                    ws.cell(row=1, column=2).value = money
                    wb.save(userlib + str(user.id) + ".xlsx")
                    wb.close()
                    embed = discord.Embed(title="KET", description=str(money) + "<:ket:753449741186105375> 추가 완료",
                                          color=0xeff0f1)
                    await ctx.send(embed=embed)
                else:
                    embed = discord.Embed(title="NO", description="유저가 ``케테르 경제``에 참여하지 않았어요..", color=0xeff0f1)
                    embed.set_thumbnail(
                        url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                    await ctx.send(embed=embed)

    @commands.command(aliases=['회사등록'])
    @commands.check(permissions.is_owner)
    async def 상장(self, ctx, name: str, stocks: int, price: int, sales: int, ratio: float):
        name = name.replace("_", " ")
        if os.path.isfile(stocklib + name + ".xlsx"):
            embed = discord.Embed(title="KMF", description="이미 상장된 기업입니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = str(int(stocks))  # 최대주
        ws.cell(row=1, column=2).value = "0"  # 매매된 주
        ws.cell(row=1, column=3).value = "1"  # 최근 거래 위치
        ws.cell(row=1, column=4).value = str(int(sales))  # 매출
        ws.cell(row=1, column=5).value = str(float(ratio))  # 수익률
        ws.cell(row=2, column=1).value = str(int(price))  # 초기가
        for i in range(2, 100):
            ws.cell(row=2, column=i).value = str(int(price))
        ws.cell(row=2, column=100).value = "1000"  # 초기설정
        wb.save(stocklib + name + ".xlsx")
        wb.close()
        time.sleep(1)
        embed = discord.Embed(title="KMF", description=name + "사 상장 완료!", color=0xeff0f1)
        await ctx.send(embed=embed)

    @commands.command(aliases=['회사삭제'])
    @commands.check(permissions.is_owner)
    async def 상장폐지(self, ctx, name: str):
        name = name.replace("_", " ")
        if os.path.isfile(stocklib + name + ".xlsx"):
            os.remove(stocklib + name + ".xlsx")
            embed = discord.Embed(title="KMF", description="해당 기업을 상장폐지 하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        embed = discord.Embed(title="KMF", description=name + "는 없는 회사명입니다.", color=0xeff0f1)
        await ctx.send(embed=embed)

    @commands.command(aliases=['회사정보'])
    async def 회사(self, ctx, name: str):
        name = name.replace("_", " ")

        def keundon(value: int):
            value = int(value)
            if value < 0:
                return "변수는 음수값을 가질 수 없습니다."
            elif 0 <= value < 10000:
                return str(value)
            elif 10000 <= value < 100000000:
                return str(math.floor(value / 10000)) + "만 " + str(value - math.floor(value / 10000) * 10000)
            elif 100000000 <= value < 1000000000000:
                return str(math.floor(value / 100000000)) + "억 " + str(
                    math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(
                    value - math.floor(value / 10000) * 10000)
            elif 1000000000000 <= value < 10000000000000000:
                return str(math.floor(value / 1000000000000)) + "조 " + str(
                    math.floor(value / 100000000) - math.floor(value / 1000000000000) * 10000) + "억 " + str(
                    math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(
                    value - math.floor(value / 10000) * 10000)
            elif 10000000000000000 <= value < 100000000000000000000:
                return str(math.floor(value / 10000000000000000)) + "경 " + str(
                    math.floor(value / 1000000000000) - math.floor(value / 10000000000000000) * 10000) + "조 " + str(
                    math.floor(value / 100000000) - math.floor(value / 1000000000000) * 10000) + "억 " + str(
                    math.floor(value / 10000) - math.floor(value / 100000000) * 10000) + "만 " + str(
                    value - math.floor(value / 10000) * 10000)
            else:
                return "변수의 크기가 너무 큽니다."

        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            stoks = ws.cell(row=1, column=1).value
            last = ws.cell(row=1, column=3).value
            sales = ws.cell(row=1, column=4).value
            ratio = ws.cell(row=1, column=5).value
            price = ws.cell(row=2, column=int(last)).value
            if last == "1":
                prece = ws.cell(row=2, column=100).value
            else:
                prece = ws.cell(row=2, column=int(last) - 1).value
            wb.close()
            siga = keundon(int(price) * int(stoks))
            perc = round(int(price) * 100 / int(prece) - 100, 2)
            if perc > 0:
                icon = ":small_red_triangle:"
            else:
                icon = ":small_red_triangle_down:"
            embed = discord.Embed(title=name, color=0xeff0f1)
            embed.add_field(name="시가총액", value=siga + " <:ket:753449741186105375>")
            embed.add_field(name="주가",
                            value=keundon(price) + " <:ket:753449741186105375> (" + icon + str(abs(perc)) + "%)")
            embed.add_field(name="매출", value=keundon(int(sales)) + " <:ket:753449741186105375>")
            embed.add_field(name="순이익",
                            value=keundon(round(int(sales) * float(ratio) / 100)) + " <:ket:753449741186105375>")
            embed.add_field(name="예상 배당금", value=keundon(
                round(int(sales) / int(stoks) * float(ratio) / 100)) + " <:ket:753449741186105375>")
            await ctx.send(embed=embed)
        else:
            embed = discord.Embed(title="NO", description="해당 이름의 회사를 찾기 못하였습니다", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command(aliases=['회사조작'])
    @commands.check(permissions.is_owner)
    async def 주식조작(self, ctx, name: str, item: str, val: int):
        """ item 항목 : 주식총수, 주가, 매출, 수익률\n수익률의 변수 val은 10이 1%입니다. """
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
        else:
            embed = discord.Embed(title="NO", description="해당 이름의 회사를 찾지 못하였습니다", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        if item == "주식총주":
            if val <= ws.cell(row=1, column=2).value:
                embed = discord.Embed(title="NO", description="총수는 매매된 주보다 적은 수로 변경할 수 없습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
            else:
                ws.cell(row=1, column=1).value = str(val)
                wb.save(stocklib + name + ".xlsx")
                wb.close()
                embed = discord.Embed(title="KMF", description="해당 사(社)의 주식총수를 변경하였습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
        if item == "주가":
            last = ws.cell(row=1, column=3).value
            if last == "100":
                next = 1
            else:
                next = int(last) + 1
            ws.cell(row=1, column=3).value = str(next)
            ws.cell(row=2, column=next).value = str(val)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 사(社)의 주가를 변경하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        if item == "매출":
            ws.cell(row=1, column=4).value = str(val)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 사(社)의 매출을 변경하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        if item == "수익률":
            val = val / 10
            if val > 100:
                embed = discord.Embed(title="NO", description="수익률은 100(%)을 넘길 수 없습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
            if val <= 0:
                embed = discord.Embed(title="NO", description="수익률은 0(%)이하일 수 없습니다.", color=0xeff0f1)
                embed.set_thumbnail(
                    url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
                await ctx.send(embed=embed)
                return None
            ws.cell(row=1, column=5).value = str(val)
            wb.save(stocklib + name + ".xlsx")
            wb.close()
            embed = discord.Embed(title="KMF", description="해당 사(社)의 수익률을 변경하였습니다.", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)
            return None
        embed = discord.Embed(title="NO", description="잘못된 변수 : " + item, color=0xeff0f1)
        embed.set_thumbnail(
            url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
        await ctx.send(embed=embed)

    @commands.command(aliases=['상장사'])
    async def 회사목록(self, ctx, plist: int):
        corps = os.listdir(stocklib)
        embed = discord.Embed(title="KMF", color=0xeff0f1)
        for i in range(0 + 10 * (plist - 1), 10 + 10 * (plist - 1)):
            try:
                embed.add_field(name=str(i + 1), value=corps[i].replace(".xlsx", ""))
            except IndexError:
                return await ctx.send(embed=embed)
        await ctx.send(embed=embed)

    @commands.command(aliases=['주식그래프'])
    async def 주식(self, ctx, name: str):
        if os.path.isfile(stocklib + name + ".xlsx"):
            wb = openpyxl.load_workbook(stocklib + name + ".xlsx")
            ws = wb.active
            last = ws.cell(row=1, column=3).value
            if last == "100":
                prices = []
                for i in range(1, 101):
                    prices.append(ws.cell(row=2, column=i).value)
            else:
                prices = []
                for i in range(int(last) + 1, 101):
                    prices.append(ws.cell(row=2, column=i).value)
                for i in range(1, int(last) + 1):
                    prices.append(ws.cell(row=2, column=i).value)
            times = list(range(1, 101))
            stk = plt.figure(figsize=(39, 18))
            plt.title(name)
            plt.xlabel('최근 거래')
            plt.ylabel('주가')
            if prices[0] < prices[99]:
                plt.plot(times, prices, color='red')
            else:
                plt.plot(times, prices, color='blue')
            stk.savefig(str(ctx.author.id) + ".png", dpi=200)
            plt.close(stk)
            await ctx.send(file=discord.File("./" + str(ctx.author.id) + ".png"))
            os.remove(str(ctx.author.id) + '.png')
        else:
            embed = discord.Embed(title="NO", description="해당 이름의 회사를 찾지 못하였습니다", color=0xeff0f1)
            embed.set_thumbnail(
                url="https://cdn.discordapp.com/attachments/750540820842807396/752684853320745000/KETER_PRESTIGE.png")
            await ctx.send(embed=embed)

    @commands.command()
    async def 그래프(self, ctx):
        # 입력 데이터

        X = np.arange(3)

        Y_A = np.arange(1, 4)

        Y_B = np.arange(3, 0, step=-1)

        # plot 입력

        line1, = plt.plot(X, Y_A, 'k-', label='A', linewidth=1)

        line2, = plt.plot(X, Y_B, 'r--', label='B', linewidth=1)

        # X 및 Y 범위 설정

        plt.xlim(X[0], X[-1])

        plt.ylim(np.min(np.append(Y_A, Y_B)), np.max(np.append(Y_A, Y_B)))

        # 그래프의 타이틀과 x, y축 라벨링

        plt.title('title', pad=10)

        plt.xlabel('X axis', labelpad=10)

        plt.ylabel('Y axis', labelpad=10)

        # 틱설정

        plt.xticks(np.linspace(X[0], X[-1], 11))

        plt.yticks(np.linspace(np.min(np.append(Y_A, Y_B)), np.max(np.append(Y_A, Y_B)), 11))

        plt.minorticks_on()

        plt.tick_params(axis='both', which='both', direction='in', pad=8, top=True, right=True)

        # 레전드 및 그리드 설정

        plt.legend(loc='upper right', fancybox=False, edgecolor='k', framealpha=1.0)

        plt.grid(color='gray', dashes=(2, 2))
        plt.savefig('graph.png')
        await ctx.send(file=discord.File("./graph.png"))
        plt.clf()

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 전체초기화(self, ctx):
        file_list = os.listdir(userlib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        for i in range(len(file_list)):
            os.remove(userlib + file_list[i])
            await ctx.send(file_list[i] + "deleted")

    @commands.command()
    @commands.check(permissions.is_owner)
    async def 상장초기화(self, ctx):
        file_list = os.listdir(stocklib)
        file_list = [file for file in file_list if file.endswith(".xlsx")]
        for i in range(len(file_list)):
            os.remove(stocklib + file_list[i])
            await ctx.send(file_list[i] + "deleted")


def setup(bot):
    bot.add_cog(economy_ko(bot))
